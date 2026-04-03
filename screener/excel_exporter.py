import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime
import pytz

# RS Trend 顏色對照
RS_TREND_COLORS = {
    "加速上升": "0F6E56",
    "穩定維持": "1B3A5C",
    "開始衰退": "C0392B",
    "震盪":     "888888",
}

GLOBAL_REGION_MAP = {
    "美國": "美洲", "加拿大": "美洲", "巴西": "美洲", "墨西哥": "美洲",
    "英國": "歐洲", "德國": "歐洲", "法國": "歐洲", "瑞士": "歐洲",
    "西班牙": "歐洲", "義大利": "歐洲", "波蘭": "歐洲", "以色列": "歐洲",
    "日本": "亞太", "中國": "亞太", "香港": "亞太", "台灣": "亞太",
    "韓國": "亞太", "印度": "亞太", "澳洲": "亞太", "紐西蘭": "亞太",
    "新加坡": "亞太", "馬來西亞": "亞太", "印尼": "亞太", "菲律賓": "亞太",
    "泰國": "亞太", "越南": "亞太",
    "杜拜": "中東",
}


def export_to_excel(df: pd.DataFrame, output_path: str, sector_ranking: list = None, global_ranking: list = None) -> str:
    """把 Screener 結果輸出成格式化 Excel"""

    tz = pytz.timezone("Asia/Taipei")
    today = datetime.now(tz).strftime("%Y-%m-%d")

    wb = openpyxl.Workbook()

    # ── Sheet 1：完整排名 ──────────────────────────────────────
    ws = wb.active
    ws.title = f"Screener_{today}"

    headers = [
        "Rank", "Change", "Ticker", "Sector", "RS Score", "RS Trend", "RS 1w", "RS 4w", "RS 13w",
        "VCP Score", "VCP Pullbacks", "Last Pullback %", "Dist from High %",
        "Combined Score", "Price", "vs 200MA %", "ATR Contraction %", "Volume Ratio"
    ]

    df_cols = [
        "Rank", "Rank_Change_Str", "Ticker", "Sector", "RS_Score", "rs_trend", "rs_1w", "rs_4w", "rs_13w",
        "Contraction_Score", "vcp_pullback_count", "last_pullback_pct", "dist_from_high_pct",
        "Combined_Score", "Price", "vs_200MA_pct", "ATR_Contraction_pct", "Volume_Ratio_10d_60d"
    ]

    col_widths = [6, 8, 8, 12, 10, 12, 8, 8, 8, 10, 14, 14, 14, 14, 10, 10, 16, 14]

    # 標題列樣式
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 數據行
    for row_idx, row in df.iterrows():
        excel_row = row_idx + 2
        for col_idx, col in enumerate(df_cols, 1):
            val = row.get(col, None)
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")

            # 數字格式
            if col == "Price":
                cell.number_format = '$#,##0.00'
            elif col in ["RS_Score", "Contraction_Score", "Combined_Score", "rs_1w", "rs_4w", "rs_13w"]:
                cell.number_format = '0.0'
            elif col in ["vs_200MA_pct", "ATR_Contraction_pct", "last_pullback_pct", "dist_from_high_pct"]:
                cell.number_format = '0.0"%"'
            elif col == "Volume_Ratio_10d_60d":
                cell.number_format = '0.00"x"'

            # RS Trend 欄位顏色
            if col == "rs_trend" and val:
                trend_color = RS_TREND_COLORS.get(str(val), "888888")
                cell.font = Font(color=trend_color, bold=True)

            # Rank Change 欄位顏色
            if col == "Rank_Change_Str" and val:
                s = str(val)
                if s.startswith("↑"):
                    cell.font = Font(color="0F6E56", bold=True)
                elif s.startswith("↓"):
                    cell.font = Font(color="C0392B", bold=True)
                elif s == "新進":
                    cell.font = Font(color="185FA5", bold=True)
                else:
                    cell.font = Font(color="888888")

            # 交替行底色
            if excel_row % 2 == 0:
                if col == "rs_trend" and val:
                    trend_color = RS_TREND_COLORS.get(str(val), "888888")
                    cell.font = Font(color=trend_color, bold=True)
                elif col == "Rank_Change_Str" and val:
                    s = str(val)
                    if s.startswith("↑"):
                        cell.font = Font(color="0F6E56", bold=True)
                    elif s.startswith("↓"):
                        cell.font = Font(color="C0392B", bold=True)
                    elif s == "新進":
                        cell.font = Font(color="185FA5", bold=True)
                cell.fill = PatternFill(start_color="F5F8FC", end_color="F5F8FC", fill_type="solid")

    # 條件格式：RS Score（C欄）綠紅漸層
    last_row = len(df) + 1
    ws.conditional_formatting.add(
        f"C2:C{last_row}",
        ColorScaleRule(
            start_type="min", start_color="C0392B",
            mid_type="percentile", mid_value=50, mid_color="FFFFFF",
            end_type="max", end_color="0F6E56",
        )
    )

    # 條件格式：Combined Score（L欄）
    ws.conditional_formatting.add(
        f"L2:L{last_row}",
        ColorScaleRule(
            start_type="min", start_color="C0392B",
            mid_type="percentile", mid_value=50, mid_color="FFFFFF",
            end_type="max", end_color="0F6E56",
        )
    )

    # 欄寬
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 凍結首行
    ws.freeze_panes = "A2"

    # ── Sheet 2：Top 30 精選（含基本面）────────────────────────
    ws2 = wb.create_sheet("Top 30")

    top30_headers = list(headers) + [
        "EPS TTM", "EPS Fwd", "EPS CAGR 2Y%", "FCF Margin%", "ROIC%", "Op Margin%", "Rev Growth%"
    ]
    top30_df_cols = list(df_cols) + [
        "eps_ttm", "eps_fwd", "eps_cagr_2y", "fcf_margin", "roic", "op_margin", "rev_growth"
    ]
    top30_widths = list(col_widths) + [10, 10, 14, 12, 10, 12, 12]

    top30_fill = PatternFill(start_color="0F6E56", end_color="0F6E56", fill_type="solid")

    for col_idx, header in enumerate(top30_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.fill = top30_fill
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in df.head(30).iterrows():
        excel_row = row_idx + 2
        for col_idx, col in enumerate(top30_df_cols, 1):
            val = row.get(col, None)
            cell = ws2.cell(row=excel_row, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            if col == "Price":
                cell.number_format = '$#,##0.00'
            elif col in ["RS_Score", "Contraction_Score", "Combined_Score", "rs_1w", "rs_4w", "rs_13w"]:
                cell.number_format = '0.0'
            elif col in ["vs_200MA_pct", "ATR_Contraction_pct", "last_pullback_pct", "dist_from_high_pct",
                          "eps_cagr_2y", "fcf_margin", "roic", "op_margin", "rev_growth"]:
                cell.number_format = '0.0"%"'
            elif col in ["eps_ttm", "eps_fwd"]:
                cell.number_format = '0.00'
            elif col == "Volume_Ratio_10d_60d":
                cell.number_format = '0.00"x"'
            if col == "rs_trend" and val:
                trend_color = RS_TREND_COLORS.get(str(val), "888888")
                cell.font = Font(color=trend_color, bold=True)
            # ROIC / FCF 填色
            if col in ("roic", "fcf_margin") and val is not None:
                try:
                    v = float(val)
                    if v > 20:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif v >= 10:
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    elif v < 0:
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                except (ValueError, TypeError):
                    pass

    for i, width in enumerate(top30_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    ws2.freeze_panes = "A2"

    # ── Sheet 3：美股類股 RS ─────────────────────────────────────
    if sector_ranking:
        ws_sec = wb.create_sheet("美股類股 RS")
        sec_headers = ["排名", "名稱", "Ticker", "RS Score", "RS Trend", "RS 1w", "RS 4w", "RS 13w", "vs SPY 13w%", "股價"]
        sec_widths = [6, 12, 8, 10, 12, 8, 8, 8, 12, 10]

        sec_fill = PatternFill(start_color="7F77DD", end_color="7F77DD", fill_type="solid")
        for col_idx, h in enumerate(sec_headers, 1):
            cell = ws_sec.cell(row=1, column=col_idx, value=h)
            cell.fill = sec_fill
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center")

        for row_idx, item in enumerate(sector_ranking, 2):
            vals = [
                item.get("rank"), item.get("name"), item.get("ticker"),
                item.get("rs_score"), item.get("rs_trend"),
                item.get("rs_1w"), item.get("rs_4w"), item.get("rs_13w"),
                item.get("vs_benchmark"), item.get("price"),
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws_sec.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="center")
                if col_idx == 4:  # RS Score
                    cell.number_format = '0.0'
                elif col_idx == 10:  # 股價
                    cell.number_format = '$#,##0.00'
                elif col_idx in (6, 7, 8):  # RS 1w/4w/13w
                    cell.number_format = '0.0'
                elif col_idx == 9:  # vs SPY
                    cell.number_format = '0.00"%"'
                if col_idx == 5 and val:  # RS Trend
                    trend_color = RS_TREND_COLORS.get(str(val), "888888")
                    cell.font = Font(color=trend_color, bold=True)
                if row_idx % 2 == 1:
                    cell.fill = PatternFill(start_color="F5F5FC", end_color="F5F5FC", fill_type="solid")

        sec_last = len(sector_ranking) + 1
        ws_sec.conditional_formatting.add(
            f"D2:D{sec_last}",
            ColorScaleRule(
                start_type="min", start_color="C0392B",
                mid_type="percentile", mid_value=50, mid_color="FFFFFF",
                end_type="max", end_color="0F6E56",
            )
        )
        for i, w in enumerate(sec_widths, 1):
            ws_sec.column_dimensions[get_column_letter(i)].width = w
        ws_sec.freeze_panes = "A2"

    # ── Sheet 4：全球指數 RS ──────────────────────────────────────
    if global_ranking:
        ws_glb = wb.create_sheet("全球指數 RS")
        glb_headers = ["排名", "地區", "市場", "Ticker", "RS Score", "RS Trend", "RS 1w", "RS 4w", "RS 13w", "vs VT 13w%"]
        glb_widths = [6, 8, 10, 12, 10, 12, 8, 8, 8, 12]

        glb_fill = PatternFill(start_color="085041", end_color="085041", fill_type="solid")
        for col_idx, h in enumerate(glb_headers, 1):
            cell = ws_glb.cell(row=1, column=col_idx, value=h)
            cell.fill = glb_fill
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center")

        for row_idx, item in enumerate(global_ranking, 2):
            region = GLOBAL_REGION_MAP.get(item.get("name", ""), "其他")
            vals = [
                item.get("rank"), region, item.get("name"), item.get("ticker"),
                item.get("rs_score"), item.get("rs_trend"),
                item.get("rs_1w"), item.get("rs_4w"), item.get("rs_13w"),
                item.get("vs_benchmark"),
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws_glb.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="center")
                if col_idx == 5:  # RS Score
                    cell.number_format = '0.0'
                elif col_idx in (7, 8, 9):  # RS 1w/4w/13w
                    cell.number_format = '0.0'
                elif col_idx == 10:  # vs VT
                    cell.number_format = '0.00"%"'
                if col_idx == 6 and val:  # RS Trend
                    trend_color = RS_TREND_COLORS.get(str(val), "888888")
                    cell.font = Font(color=trend_color, bold=True)
                if row_idx % 2 == 1:
                    cell.fill = PatternFill(start_color="E8F0ED", end_color="E8F0ED", fill_type="solid")

        glb_last = len(global_ranking) + 1
        ws_glb.conditional_formatting.add(
            f"E2:E{glb_last}",
            ColorScaleRule(
                start_type="min", start_color="C0392B",
                mid_type="percentile", mid_value=50, mid_color="FFFFFF",
                end_type="max", end_color="0F6E56",
            )
        )
        for i, w in enumerate(glb_widths, 1):
            ws_glb.column_dimensions[get_column_letter(i)].width = w
        ws_glb.freeze_panes = "A2"

    # ── Sheet 說明 ─────────────────────────────────────────────
    ws3 = wb.create_sheet("說明")

    explanation = [
        ("RS Score（相對強度）", True),
        ("- RS 1w/4w/13w：個股在各時間段的漲跌幅百分位排名（0-100）", False),
        ("- RS Trend：加速上升 = 短期RS > 中期RS > 長期RS（最強訊號）", False),
        ("- RS Score = 加權平均 + 趨勢加分/扣分", False),
        ("", False),
        ("VCP Score（價格收縮形態）", True),
        ("- VCP Pullbacks：近90日內的回撤次數（2-4次最理想）", False),
        ("- Last Pullback %：最後一次回撤幅度（越小越好，< 5% 為佳）", False),
        ("- Dist from High %：目前距前期高點距離（越小越接近突破點）", False),
        ("- ATR Contraction %：近10日ATR vs 近60日ATR的收縮幅度", False),
        ("", False),
        ("Combined Score = RS Score × 60% + VCP Score × 40%", True),
        ("", False),
        ("理想的買點特徵：", True),
        ("RS Score > 80 + RS Trend = 加速上升", False),
        ("VCP Pullbacks = 2-3次", False),
        ("Last Pullback % < 5%", False),
        ("Dist from High % < 3%", False),
        ("Volume Ratio < 0.8（量能萎縮）", False),
    ]

    title_font = Font(bold=True, size=12, color="1B3A5C")
    body_font = Font(size=11, color="333333")

    for row_idx, (text, is_title) in enumerate(explanation, 1):
        cell = ws3.cell(row=row_idx, column=1, value=text)
        cell.font = title_font if is_title else body_font

    ws3.column_dimensions["A"].width = 60

    wb.save(output_path)
    print(f"  ✓ Excel 輸出：{output_path}")
    return output_path
